"""Task-scoped storage and human-readable audit export for factor mining."""
from __future__ import annotations

import hashlib
import itertools
import json
import os
import platform
import sqlite3
import subprocess
import sys
import uuid
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Mapping

import pandas as pd


_TABLES = ("search_progress", "window_results", "candidate_summary", "winners", "errors")


def _json_default(value: Any) -> str:
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.isoformat()
    if hasattr(value, "item"):
        return value.item()
    return str(value)


def _json_payload(value: Mapping[str, Any]) -> str:
    return json.dumps(dict(value), ensure_ascii=False, default=_json_default, allow_nan=True)


def _safe_factor_id(value: str) -> str:
    result = "".join(char if char.isalnum() or char in {"-", "_"} else "_" for char in str(value))
    return result.strip("._") or "factor"


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _git_state(path: Path) -> dict[str, Any]:
    try:
        root = subprocess.run(
            ["git", "-C", str(path.parent), "rev-parse", "--show-toplevel"],
            check=True, capture_output=True, text=True, timeout=5,
        ).stdout.strip()
        commit = subprocess.run(
            ["git", "-C", root, "rev-parse", "HEAD"],
            check=True, capture_output=True, text=True, timeout=5,
        ).stdout.strip()
        status = subprocess.run(
            ["git", "-C", root, "status", "--short"],
            check=True, capture_output=True, text=True, timeout=5,
        ).stdout.splitlines()
        return {"root": root, "commit": commit, "dirty": bool(status), "changes": status}
    except Exception as exc:
        return {"available": False, "error": f"{type(exc).__name__}: {exc}"}


def _atomic_yaml(path: Path, payload: Mapping[str, Any]) -> None:
    import yaml

    temporary = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
    temporary.write_text(yaml.safe_dump(dict(payload), allow_unicode=True, sort_keys=False), encoding="utf-8")
    os.replace(temporary, path)


@dataclass
class ResultStore:
    path: Path
    connection: sqlite3.Connection = field(init=False)

    def __post_init__(self) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.connection = sqlite3.connect(self.path)
        self.connection.execute("PRAGMA journal_mode=WAL")
        self.connection.execute("PRAGMA synchronous=NORMAL")
        for table in _TABLES:
            self.connection.execute(
                f"CREATE TABLE IF NOT EXISTS {table} ("
                "id INTEGER PRIMARY KEY AUTOINCREMENT, event TEXT, stage TEXT, factor_id TEXT, "
                "candidate_id TEXT, window_id TEXT, payload_json TEXT NOT NULL)"
            )
            self.connection.execute(
                f"CREATE INDEX IF NOT EXISTS idx_{table}_candidate ON {table}(stage, candidate_id)"
            )
        self.connection.commit()

    def append(self, table: str, rows: Iterable[Mapping[str, Any]]) -> int:
        if table not in _TABLES:
            raise ValueError(f"unknown mining result table: {table}")
        payloads = [(
            row.get("event"), row.get("stage"), row.get("factor_id"),
            row.get("candidate_id"), row.get("window_id"), _json_payload(row),
        ) for row in rows]
        if not payloads:
            return 0
        self.connection.executemany(
            f"INSERT INTO {table}(event, stage, factor_id, candidate_id, window_id, payload_json) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            payloads,
        )
        self.connection.commit()
        return len(payloads)

    def read(self, table: str) -> pd.DataFrame:
        if table not in _TABLES:
            raise ValueError(f"unknown mining result table: {table}")
        rows = self.connection.execute(f"SELECT payload_json FROM {table} ORDER BY id").fetchall()
        return pd.DataFrame([json.loads(row[0]) for row in rows])

    def close(self) -> None:
        self.connection.commit()
        self.connection.close()


def _empty_safe(frame: pd.DataFrame, fallback: str = "暂无记录") -> pd.DataFrame:
    if not frame.empty:
        return frame
    return pd.DataFrame({"说明": [fallback]})


_COLUMN_NAMES = {
    "event": "记录类型", "stage": "搜索阶段", "factor_id": "因子",
    "candidate_id": "候选编号", "candidate_order": "候选序号",
    "completed_order": "完成序号", "trial_number": "试验编号",
    "source": "候选来源", "params_json": "参数JSON", "anchors_json": "粗搜锚点",
    "local_bounds_json": "细搜局部边界", "window_id": "窗口编号",
    "window_start": "窗口开始", "window_end": "窗口结束", "ann_ret": "年化收益率",
    "ann_vol": "年化波动率", "sharpe": "夏普比率", "mdd": "最大回撤",
    "calmar": "卡玛比率", "turnover": "换手率", "rank_ic": "Rank IC",
    "ic_coverage": "IC覆盖率", "n_days": "有效天数", "error": "错误信息",
    "error_type": "错误类型", "valid_windows": "有效窗口数",
    "failed_windows": "失败窗口数", "candidate_elapsed_seconds": "候选耗时（秒）",
    "sharpe_median": "夏普中位数", "rank": "排名", "selection_status": "筛选状态",
    "selection_reason": "筛选说明", "objective": "目标值", "valid_window_ratio": "有效窗口比例",
    "valid_window_count": "有效窗口数", "window_count": "窗口总数", "max_mdd": "最大回撤",
    "boundary_pressure_json": "边界压力", "previous_bounds_json": "扩展前边界",
    "expanded_bounds_json": "扩展后边界", "parent_candidate_id": "所属赢家候选",
    "winner_rank": "赢家排名", "perturbation_index": "扰动序号",
    "stability_status": "稳定性结论", "perturbation_count": "计划扰动数",
    "valid_perturbation_count": "有效扰动数", "valid_perturbation_ratio": "有效扰动比例",
    "perturbation_pass_ratio": "扰动通过比例", "perturbed_objective_median": "扰动目标中位数",
    "objective_degradation_median": "目标退化中位数",
}


def _humanize_columns(frame: pd.DataFrame) -> pd.DataFrame:
    suffixes = {"mean": "均值", "median": "中位数", "p25": "25%分位数", "min": "最小值", "max": "最大值"}
    metric_names = {
        "ann_ret": "年化收益率", "ann_vol": "年化波动率", "sharpe": "夏普比率",
        "mdd": "最大回撤", "calmar": "卡玛比率", "turnover": "换手率",
        "rank_ic": "Rank IC", "ic_coverage": "IC覆盖率", "n_days": "有效天数",
    }
    names = {}
    for column in frame.columns:
        translated = _COLUMN_NAMES.get(str(column))
        if translated is None:
            for suffix, suffix_name in suffixes.items():
                marker = f"_{suffix}"
                if str(column).endswith(marker):
                    metric = str(column)[:-len(marker)]
                    translated = f"{metric_names.get(metric, metric)}{suffix_name}"
                    break
        names[column] = translated or str(column)
    output = frame.rename(columns=names)
    value_mappings = {
        "搜索阶段": {
            "coarse": "宽范围粗搜",
            "refine": "自适应收敛搜索",
            "fine": "局部网格细搜",
            "stability": "赢家扰动验证",
        },
        "记录类型": {
            "planned": "已规划",
            "completed": "已完成",
            "summarized": "已汇总",
            "boundary_check": "边界检查",
        },
        "筛选状态": {
            "selected": "已入选",
            "filtered": "被过滤",
            "candidate": "候选",
        },
        "稳定性结论": {
            "stable": "稳定",
            "unstable": "不稳定",
            "not_tested": "未验证",
        },
    }
    for column, mapping in value_mappings.items():
        if column in output:
            output[column] = output[column].replace(mapping)
    if "搜索阶段" in output:
        output["搜索阶段"] = output["搜索阶段"].map(
            lambda value: f"第{str(value).rsplit('_', 1)[-1]}轮边界扩展搜索"
            if str(value).startswith("expansion_") else value
        )
    return output


def _write_sheet(writer: pd.ExcelWriter, name: str, frame: pd.DataFrame, *, max_rows: int = 1_000_000) -> None:
    frame = _empty_safe(frame)
    if name == "窗口表现" and len(frame) > max_rows:
        for offset in range(0, len(frame), max_rows):
            frame.iloc[offset:offset + max_rows].to_excel(
                writer,
                sheet_name=f"窗口表现_{offset // max_rows + 1}",
                index=False,
            )
    else:
        frame.to_excel(writer, sheet_name=name[:31], index=False)


def _style_workbook(path: Path) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill

    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for column in sheet.columns:
            values = [str(cell.value or "") for cell in list(column)[:200]]
            width = min(42, max(10, max((len(value) for value in values), default=10) + 2))
            sheet.column_dimensions[column[0].column_letter].width = width
            header = str(column[0].value or "")
            if any(word in header for word in ("收益率", "波动率", "回撤", "换手率", "覆盖率", "比例")):
                for cell in list(column)[1:]:
                    cell.number_format = "0.00%"
    workbook.save(path)


def _total_range(metadata: Mapping[str, Any]) -> tuple[pd.Timestamp, pd.Timestamp] | None:
    """Resolve the complete evaluation range used by audit heatmaps."""
    configuration = metadata.get("configuration", {})
    parameter_space = configuration.get("parameter_space", {}) if isinstance(configuration, Mapping) else {}
    evaluation = parameter_space.get("evaluation", {}) if isinstance(parameter_space, Mapping) else {}
    span = evaluation.get("span") if isinstance(evaluation, Mapping) else None
    if not span and isinstance(parameter_space, Mapping):
        span = parameter_space.get("span")
    if not isinstance(span, (list, tuple)) or len(span) != 2:
        return None
    start, end = pd.Timestamp(span[0]).normalize(), pd.Timestamp(span[1]).normalize()
    if start > end:
        return None
    return start, end


def _heatmap_parameter_pairs(frame: pd.DataFrame, parameter_specs: Mapping[str, Any]) -> list[tuple[str, str]]:
    varying = []
    for name in parameter_specs:
        if name not in frame:
            continue
        values = frame[name].dropna()
        if values.nunique(dropna=True) > 1:
            varying.append(name)
    if len(varying) >= 2:
        return list(itertools.combinations(varying, 2))
    if len(varying) == 1:
        return [(varying[0], "候选")]
    return [("候选", "候选")]


def _heatmap_matrix(frame: pd.DataFrame, row_name: str, column_name: str, metric: str) -> pd.DataFrame:
    work = frame.copy()
    if row_name == "候选":
        work[row_name] = work["candidate_id"].astype(str)
    if column_name == "候选":
        work[column_name] = work["candidate_id"].astype(str)
    work[metric] = pd.to_numeric(work.get(metric), errors="coerce")
    work = work.dropna(subset=[row_name, column_name, metric])
    if work.empty:
        return pd.DataFrame()
    # Use the positional level here.  The fallback layout can intentionally
    # use the same label ("候选") for both axes, and unstack("候选") is
    # ambiguous when pandas sees two levels with that name.
    grouped = work.groupby([row_name, column_name], dropna=False)[metric].mean()
    matrix = grouped.unstack(level=1).sort_index()
    matrix.columns.name = column_name
    return matrix


def _write_heatmap_report(
    audit_dir: Path,
    metadata: Mapping[str, Any],
    window_results: pd.DataFrame,
) -> list[str]:
    """Write parameter heatmap PNGs and a machine-readable chart index."""
    if window_results.empty:
        return []
    valid = window_results.loc[window_results.get("error", pd.Series(index=window_results.index)).isna()].copy()
    valid["window_start"] = pd.to_datetime(valid.get("window_start"), errors="coerce").dt.normalize()
    valid["window_end"] = pd.to_datetime(valid.get("window_end"), errors="coerce").dt.normalize()
    valid = valid.dropna(subset=["window_start", "window_end"])
    total_range = _total_range(metadata)
    if total_range is None:
        return []
    period_start, period_end = total_range
    total_mask = (valid["window_end"] >= period_start) & (valid["window_start"] <= period_end)
    total_frame = valid.loc[total_mask].copy()
    window_parts = total_frame.get("window_id", pd.Series(index=total_frame.index, dtype=object)).astype(str).str.extract(
        r"^(?P<window_length>\d+)/(?P<window_step>\d+)(?:/|$)"
    )
    total_frame["_window_length"] = pd.to_numeric(window_parts["window_length"], errors="coerce")
    total_frame["_window_step"] = pd.to_numeric(window_parts["window_step"], errors="coerce")
    total_frame = total_frame.dropna(subset=["_window_length", "_window_step"])
    total_frame["_window_length"] = total_frame["_window_length"].astype(int)
    total_frame["_window_step"] = total_frame["_window_step"].astype(int)
    parameter_specs = metadata.get("resolved_parameter_specs", {})
    if not isinstance(parameter_specs, Mapping):
        parameter_specs = {}
    pairs = _heatmap_parameter_pairs(total_frame, parameter_specs)
    metrics = [("sharpe", "夏普比率"), ("ann_ret", "年化收益率"), ("calmar", "卡玛比率"), ("mdd", "最大回撤")]
    output_paths: list[str] = []
    index_rows: list[dict[str, Any]] = []
    try:
        import matplotlib

        matplotlib.use("Agg", force=True)
        import matplotlib.pyplot as plt
        import numpy as np
        from matplotlib import font_manager

        # Prefer an installed CJK font so Chinese worksheet/report labels are
        # rendered instead of producing glyph-missing warnings or tofu boxes.
        preferred_fonts = ("Microsoft YaHei", "SimHei", "Noto Sans CJK SC", "Arial Unicode MS")
        available_fonts = {font.name for font in font_manager.fontManager.ttflist}
        for font_name in preferred_fonts:
            if font_name in available_fonts:
                plt.rcParams["font.sans-serif"] = [font_name]
                break
        plt.rcParams["axes.unicode_minus"] = False
    except ImportError as exc:
        index_path = audit_dir / "热力图报告.txt"
        index_path.write_text(f"无法生成热力图：缺少 matplotlib（{exc}）\n", encoding="utf-8")
        return [str(index_path)]
    audit_dir.mkdir(parents=True, exist_ok=True)
    schemes = total_frame[["_window_length", "_window_step"]].drop_duplicates().sort_values(
        ["_window_length", "_window_step"]
    )
    for window_length, window_step in schemes.itertuples(index=False, name=None):
        scheme_frame = total_frame.loc[
            total_frame["_window_length"].eq(window_length)
            & total_frame["_window_step"].eq(window_step)
        ]
        for pair_index, (row_name, column_name) in enumerate(pairs, 1):
            figure, axes = plt.subplots(2, 2, figsize=(16, 11), constrained_layout=True)
            figure.suptitle(
                f"{metadata.get('factor_id', '因子')} 参数表现热力图 · total\n"
                f"{window_length}日窗口 / 每{window_step}日滑动 · "
                f"参数轴：{row_name} × {column_name} · 聚合：平均值",
                fontsize=14,
            )
            for axis, (metric, title) in zip(axes.ravel(), metrics):
                matrix = _heatmap_matrix(scheme_frame, row_name, column_name, metric)
                if matrix.empty:
                    axis.text(0.5, 0.5, "暂无有效结果", ha="center", va="center", fontsize=13)
                    axis.set_axis_off()
                    continue
                values = matrix.to_numpy(dtype=float)
                image = axis.imshow(values, aspect="auto", cmap="RdYlGn", interpolation="nearest")
                axis.set_title(title)
                axis.set_xlabel(column_name)
                axis.set_ylabel(row_name)
                axis.set_xticks(range(len(matrix.columns)), [str(value) for value in matrix.columns], rotation=45, ha="right")
                axis.set_yticks(range(len(matrix.index)), [str(value) for value in matrix.index])
                if values.size <= 144:
                    finite = values[np.isfinite(values)]
                    threshold = (float(finite.min()) + float(finite.max())) / 2 if finite.size else 0.0
                    for row_index in range(values.shape[0]):
                        for column_index in range(values.shape[1]):
                            value = values[row_index, column_index]
                            if np.isfinite(value):
                                axis.text(column_index, row_index, f"{value:.3g}", ha="center", va="center", fontsize=7, color="black" if value >= threshold else "white")
                figure.colorbar(image, ax=axis, shrink=0.82)
            path = audit_dir / (
                f"热力图_total_{window_length}日窗口_{window_step}日步长_参数对{pair_index:02d}.png"
            )
            figure.savefig(path, dpi=150, bbox_inches="tight")
            plt.close(figure)
            output_paths.append(str(path))
            index_rows.append({
                "period": "total",
                "start": period_start.date().isoformat(),
                "end": period_end.date().isoformat(),
                "window_length": int(window_length),
                "window_step": int(window_step),
                "window_count": int(scheme_frame["window_id"].nunique()),
                "rows": int(len(scheme_frame)),
                "aggregation": "mean",
                "search_stages": sorted(scheme_frame.get("stage", pd.Series(dtype=str)).dropna().astype(str).unique()),
                "row_parameter": row_name,
                "column_parameter": column_name,
                "path": str(path),
            })
    report_path = audit_dir / "热力图报告.json"
    report_path.write_text(json.dumps(index_rows, ensure_ascii=False, indent=2), encoding="utf-8")
    output_paths.append(str(report_path))
    return output_paths


@dataclass
class FactorMiningResult:
    factor_id: str
    run_id: str
    run_dir: Path
    status: str
    coarse_window_results: pd.DataFrame = field(default_factory=pd.DataFrame)
    coarse_summary: pd.DataFrame = field(default_factory=pd.DataFrame)
    fine_window_results: pd.DataFrame = field(default_factory=pd.DataFrame)
    fine_summary: pd.DataFrame = field(default_factory=pd.DataFrame)
    stability_window_results: pd.DataFrame = field(default_factory=pd.DataFrame)
    stability_summary: pd.DataFrame = field(default_factory=pd.DataFrame)
    stage_window_results: dict[str, pd.DataFrame] = field(default_factory=dict)
    stage_summaries: dict[str, pd.DataFrame] = field(default_factory=dict)
    selected_candidates: pd.DataFrame = field(default_factory=pd.DataFrame)
    heatmap_paths: tuple[str, ...] = ()


@dataclass
class MiningResult:
    launch_id: str
    factor_runs: tuple[FactorMiningResult, ...]


@dataclass
class MiningTask:
    factor_id: str
    run_id: str
    launch_id: str
    run_dir: Path
    cache_dir: Path
    audit_dir: Path
    metadata_path: Path
    workbook_path: Path
    log_path: Path
    store: ResultStore
    metadata: dict[str, Any]

    @classmethod
    def create(
        cls,
        output_root: str | Path,
        factor_id: str,
        launch_id: str,
        *,
        factor_class: str | None,
        parameter_path: Path,
        performance_path: Path,
        config: Mapping[str, Any],
    ) -> "MiningTask":
        now = datetime.now().astimezone()
        git = _git_state(parameter_path)
        run_id = f"{now.strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}"
        factor_dir = Path(output_root).resolve() / _safe_factor_id(factor_id)
        run_dir = factor_dir / run_id
        run_dir.mkdir(parents=True, exist_ok=False)
        cache_dir = run_dir / "cache"
        audit_dir = run_dir / "audit"
        cache_dir.mkdir()
        audit_dir.mkdir()
        metadata_path = run_dir / "metadata.yaml"
        task = cls(
            factor_id=str(factor_id),
            run_id=run_id,
            launch_id=launch_id,
            run_dir=run_dir,
            cache_dir=cache_dir,
            audit_dir=audit_dir,
            metadata_path=metadata_path,
            workbook_path=audit_dir / "挖掘审计.xlsx",
            log_path=audit_dir / "运行日志.log",
            store=ResultStore(cache_dir / "results.sqlite3"),
            metadata={
                "schema_version": 1,
                "status": "running",
                "launch_id": launch_id,
                "run_id": run_id,
                "factor_id": str(factor_id),
                "factor_class": factor_class,
                "started_at_local": now.isoformat(),
                "started_at_utc": now.astimezone(timezone.utc).isoformat(),
                "parameter_config": {"path": str(parameter_path), "sha256": _sha256(parameter_path)},
                "performance_config": {"path": str(performance_path), "sha256": _sha256(performance_path)},
                "configuration": dict(config),
                "environment": {
                    "python": sys.version,
                    "platform": platform.platform(),
                    "hostname": platform.node(),
                    "pid": os.getpid(),
                },
                "git": git,
                "paths": {
                    "run_dir": str(run_dir),
                    "cache_dir": str(cache_dir),
                    "audit_dir": str(audit_dir),
                    "workbook": str(audit_dir / "挖掘审计.xlsx"),
                    "log": str(audit_dir / "运行日志.log"),
                    "results_db": str(cache_dir / "results.sqlite3"),
                },
            },
        )
        task.write_metadata()
        return task

    def write_metadata(self, **updates: Any) -> None:
        self.metadata.update(updates)
        _atomic_yaml(self.metadata_path, self.metadata)

    def export_workbook(self) -> Path:
        overview = pd.DataFrame(
            [
                {
                    "项目": key,
                    "内容": json.dumps(value, ensure_ascii=False, default=_json_default)
                    if isinstance(value, (dict, list)) else value,
                }
                for key, value in self.metadata.items()
                if key != "configuration"
            ]
        )
        config_rows = []
        for section, values in self.metadata.get("configuration", {}).items():
            if isinstance(values, Mapping):
                for key, value in values.items():
                    config_rows.append({"配置段": section, "配置项": key, "配置值": json.dumps(value, ensure_ascii=False, default=_json_default) if isinstance(value, (dict, list)) else value})
            else:
                config_rows.append({"配置段": "", "配置项": section, "配置值": values})
        parameter_rows = []
        parameter_config = self.metadata.get("configuration", {}).get("parameter_space", {})
        for factor_id, factor in (parameter_config.get("factors") or {}).items() if isinstance(parameter_config, Mapping) else []:
            if str(factor_id) != self.factor_id:
                continue
            for name, spec in (factor.get("parameters") or {}).items():
                row = {"因子": factor_id, "参数": name, **dict(spec)}
                parameter_rows.append(row)
        if not parameter_rows:
            for name, spec in (self.metadata.get("resolved_parameter_specs") or {}).items():
                parameter_rows.append({"因子": self.factor_id, "参数": name, **dict(spec)})
        progress = self.store.read("search_progress")
        windows = self.store.read("window_results")
        summaries = self.store.read("candidate_summary")
        winners = self.store.read("winners")
        errors = self.store.read("errors")
        stability_details = summaries.loc[
            summaries.get("stage", pd.Series(index=summaries.index, dtype=object)).eq("stability")
        ] if not summaries.empty else pd.DataFrame()
        winner_summaries = winners.copy()
        stability_parts = []
        if not winners.empty and "stability_status" in winners:
            conclusions = winners.copy()
            conclusions.insert(0, "record_kind", "赢家结论")
            stability_parts.append(conclusions)
        if not stability_details.empty:
            details = stability_details.copy()
            details.insert(0, "record_kind", "扰动候选")
            stability_parts.append(details)
        stability = pd.concat(stability_parts, ignore_index=True, sort=False) if stability_parts else pd.DataFrame()
        winner_parameter_columns = [
            name for name in ("rank", "factor_id", "candidate_id", "params_json", "objective")
            if name in winners
        ]
        parameter_names = set()
        if "params_json" in winners:
            for payload in winners["params_json"].dropna():
                try:
                    parameter_names.update(json.loads(str(payload)))
                except (TypeError, ValueError):
                    pass
        winner_parameter_columns.extend(
            name for name in winners.columns
            if name in parameter_names and name not in winner_parameter_columns
        )
        winner_parameters = winners.loc[:, winner_parameter_columns] if winner_parameter_columns else winners
        self.workbook_path.parent.mkdir(parents=True, exist_ok=True)
        temporary = self.workbook_path.with_name(f".{self.workbook_path.name}.{uuid.uuid4().hex}.tmp.xlsx")
        with pd.ExcelWriter(temporary, engine="openpyxl") as writer:
            _write_sheet(writer, "运行概览", _humanize_columns(overview))
            _write_sheet(writer, "运行配置", _humanize_columns(pd.DataFrame(config_rows)))
            _write_sheet(writer, "参数空间", _humanize_columns(pd.DataFrame(parameter_rows)))
            _write_sheet(writer, "搜索进度", _humanize_columns(progress))
            _write_sheet(writer, "窗口表现", _humanize_columns(windows))
            _write_sheet(writer, "候选汇总", _humanize_columns(summaries))
            _write_sheet(writer, "赢家参数", _humanize_columns(winner_parameters))
            _write_sheet(writer, "赢家汇总", _humanize_columns(winner_summaries))
            _write_sheet(writer, "稳定性验证", _humanize_columns(stability))
            _write_sheet(writer, "错误", _humanize_columns(errors))
        _style_workbook(temporary)
        os.replace(temporary, self.workbook_path)
        return self.workbook_path

    def finish(self, result: FactorMiningResult, *, status: str, **updates: Any) -> None:
        windows = self.store.read("window_results")
        summaries = self.store.read("candidate_summary")
        winners = self.store.read("winners")
        errors = self.store.read("errors")
        window_stages = windows.get("stage", pd.Series(index=windows.index, dtype=object))
        summary_stages = summaries.get("stage", pd.Series(index=summaries.index, dtype=object))
        stage_window_rows = {
            str(stage): int(count)
            for stage, count in window_stages.value_counts().to_dict().items()
        }
        stage_candidates = {
            str(stage): int(count)
            for stage, count in summary_stages.value_counts().to_dict().items()
        }
        self.write_metadata(
            status=status,
            completed_at_local=datetime.now().astimezone().isoformat(),
            completed_at_utc=datetime.now().astimezone(timezone.utc).isoformat(),
            result_counts={
                "coarse_window_rows": int(window_stages.eq("coarse").sum()),
                "fine_window_rows": int(window_stages.eq("fine").sum()),
                "coarse_candidates": int(summary_stages.eq("coarse").sum()),
                "fine_candidates": int(summary_stages.eq("fine").sum()),
                "selected_candidates": len(winners),
                "errors": len(errors),
                "stage_window_rows": stage_window_rows,
                "stage_candidates": stage_candidates,
            },
            winners=winners.to_dict(orient="records"),
            **updates,
        )
        # Heatmaps are an audit supplement.  A missing plotting dependency or
        # malformed optional metric must not erase the primary mining result.
        try:
            heatmap_paths = tuple(_write_heatmap_report(self.audit_dir, self.metadata, windows))
        except Exception as exc:
            heatmap_paths = ()
            result_counts = dict(self.metadata.get("result_counts") or {})
            result_counts["errors"] = int(result_counts.get("errors", 0)) + 1
            self.write_metadata(
                heatmap_error=f"{type(exc).__name__}: {exc}",
                result_counts=result_counts,
            )
            self.store.append(
                "errors",
                [{
                    "factor_id": self.factor_id,
                    "candidate_id": "",
                    "window_id": "",
                    "stage": "audit",
                    "error_type": type(exc).__name__,
                    "error": str(exc),
                }],
            )
        if heatmap_paths:
            self.write_metadata(heatmap_paths=list(heatmap_paths))
            result.heatmap_paths = heatmap_paths
        self.export_workbook()
        self.store.close()


__all__ = ["FactorMiningResult", "MiningResult", "MiningTask", "ResultStore"]
